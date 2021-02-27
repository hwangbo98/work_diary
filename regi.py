import math
import getpass
import pickle
import os
import sys
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
from datetime import datetime, timezone, date, time
from pytz import timezone, utc

def register(data) :
	
	x = input("Do you want to register? : yes or no ")
	x_lower = x.lower()
	if x_lower == 'yes':
		while(1) :
			account = input("Which ID do you want? :")
			if account in data.keys() :
				print("already exist")
			else :
				password = getpass.getpass("which password do you want? :") 
				break
		
	else :
		sys.exit()

	data[account] = password
	#wb = openpyxl.load_workbook("work_sheet.xlsx",data_only = True)
	#sheet_last = wb.create_sheet(account)

	with open('account.pkl', 'wb') as account_file :
		pickle.dump(data,account_file)
	print(account) 
	print(password)
	return data
def login(data) :
	while(1) :
		id_input = input("Write the ID :")
		pw_input = getpass.getpass("Write the PW :")

		if id_input in data :
			if data[id_input] == pw_input :
				print("Login Success")
				break
			else :
				print("Password Wrong")

		else :
			print("rewrite the ID & PW")

	return id_input
def time_service(id_info) :
	KST = timezone('Asia/Seoul') #ubuntu default UTC Time
	split_day_time =[]
	list_h_m_s = []
	now_time = datetime.utcnow()
	timeTostr = utc.localize(now_time).astimezone(KST).strftime("%Y-%m-%d %H:%M:%S")
	#timeTostr1 = utc.localize(now_time).astimezone(KST).strftime("%H:%M:%S")
	get_month = utc.localize(now_time).astimezone(KST).strftime("%B")
	strToint = datetime.strptime(timeTostr,"%Y-%m-%d %H:%M:%S")
	split_day_time = strToint.strftime("%Y-%m-%d %H:%M:%S").split()	
	print(split_day_time[0])
	print(split_day_time[1])
	split_day_time.append(id_info)
	
	split_day_time.append(get_month[:3]) #get month name
	print(split_day_time[2])
	#print(split_day_time[3])
	return split_day_time

def go_work_excel(day_time_name_info) :
	file_name = day_time_name_info[2] + "_" + day_time_name_info[3] + '.csv'
	f = open(file_name,'r', encoding = 'cp949') 
	read_file = csv.reader(f)
	lines = []
	for line in read_file :
		if line[0] == day_time_name_info[0] :
			line[1] = day_time_name_info[1]
		lines.append(line)
	f = open(file_name,'w')
	wr = csv.writer(f)
	wr.writerows(lines)

	f.close()
def leave_office_excel(day_time_name_info) :
	file_name = day_time_name_info[2] + "_" + day_time_name_info[3] + '.csv'
	f = open(file_name,'r', encoding = 'cp949')
	read_file = csv.reader(f)
	lines = []
	for line in read_file :
		if line[0] == day_time_name_info[0] :
			line[2] = day_time_name_info[1]
			go_time = line[0] + " " + line[1]
			leave_time = line[0] + " "+ line[2]
			go_strToint = datetime.strptime(go_time, "%Y-%m-%d %H:%M:%S")
			leave_strToint = datetime.strptime(leave_time, "%Y-%m-%d %H:%M:%S")
			#print(go_strToint)
			#print(leave_strToint)
			sub_time = leave_strToint - go_strToint
			#print(sub_time)
			line[3] = sub_time.total_seconds()/60
			line[3] =  round(line[3]) #trunc is waste of a decimal num
		lines.append(line)
	f = open(file_name,'w')
	wr = csv.writer(f)
	wr.writerows(lines)
	
	f.close()
def print_work_record(data) :
	while(1) :
		whose_record = input("whose record do you want to see? please type the ID : ")
		print(data.keys())
		
		if whose_record in data.keys() :
			get_month = time_service(whose_record)
			file_name = whose_record + "_" + get_month[3] + '.csv'
			f = open(file_name,'r', encoding ='cp949')
			read_file = csv.reader(f)
			for line in read_file :
				print(line)
			f.close()
			break
		else :
			print("Wrong ID. Type correctly")
def total_time(name_worker) :
	get_month = time_service(name_worker)
	file_name = name_worker + "_" + get_month[3] + '.csv'
	f = open(file_name,'r', encoding ='cp949')
	read_file = csv.reader(f)
	times = []
	count = 0
	for line in read_file :
		if count == 0 :
			pass
		else :
			if line[3] == '' :
				line[3] = 0
				times.append(line[3])
			else :
				times.append(line[3])
		count += 1

	times = list(map(int,times))
	print("name_worker work time : %d"%(sum(map(int,times))))	
	print(times)
	
'''def go_work_excel(day_time_name_info) : #it is defined go to work example
	wb = openpyxl.load_workbook("work_sheet.xlsx",data_only = True)		
	now_sheet = wb.get_sheet_by_name([day_time_name_info[2]])
	day_column = now_sheet['A']
	count1 = 0 
	for day_day in day_column :
		count1 = count1 + 1
		if day_day == day_time_name_info[0] :
			now_sheet.cell(row = 2, column = count1).value = day_time_name_info[1]
	print(count1)

	wb.save("work_sheet.xlsx")
	wb.close()
def leave_office_excel(day_time_name_info) :
	wb = openpyxl.load_workbook("work_sheet.xlsx",data_only = True)
	now_sheet = wb.get_sheet_by_name([day_time_name_info[2]])
	day_column = now_sheet['A']
	count2 = 0
	for day_day in day_column :
		count2 = count2 + 1
		if day_day == day_time_name_info[0] :
			now_sheet.cell(row = 3, column = count2).value = day_time_name_info[1]

	go_work = now_sheet.cell(row = 2, column = count2).value
	leave_office = now_sheet.cell(row = 3, column = count2).value

	worktime_of_day = leave_office - go_work 
	now_sheet.cell(row = 4, column = count2).value = worktime_of_day
	
	wb.save("work_sheet.xlsx")
	wb.close()
'''	
id_pw = dict()
day_time_name = []
filesize = os.path.getsize('account.pkl')
if filesize == 0 :
	pass
else :	
	with open('account.pkl', 'rb') as fin :
		id_pw = pickle.load(fin)
print("You must choice the number, What do you want? ")
#print(id_pw.keys())
print("If you want to quit or wrong something, type quit() and you will be OK ")
choice = input("1.register 2.login 3.logout " )
if choice == "1" :
	id_pw = register(id_pw)
elif choice == "2" :
	who_am_i = login(id_pw)
	if who_am_i == 'admin' :
		answer_admin = int(input("1. delete member 2. print out the member's monthly work record. 3. total time of specific member?  "))
		if answer_admin == 1 :
			while(1) :
				name_delete = input("which account do you want to delete? ")
				if name_delete not in id_pw.keys() or name_delete == 'admin' :
					print("incorrect ID. You type the correct ID one more!")
				else :
					admin_pw = getpass.getpass("type the admin's password ")
					yes_or_no = input("if you delete the member, you cannot revive the member yes or no? ")
					yes_or_no = yes_or_no.lower()
					if id_pw[who_am_i] == admin_pw and yes_or_no == 'yes' :

						del(id_pw[name_delete])
						print("delete account succesful")
						with open('account.pkl', 'wb') as account_file :
							pickle.dump(id_pw,account_file)
						break
					else :
						print("Wrong Password or you don't want to delete")
					
		elif answer_admin == 2 :
			print_work_record(id_pw)
		elif answer_admin == 3 :
			while(1) :
				time_name = input("type the name that what you want to see the monthly total work time ")
				if time_name in id_pw.keys() :
					total_time(time_name)
					break
				else :
					print("type one more time! {} doesn't exist".format(time_name))

	else :
		answer = int(input("1. go to work 2. leave the office :"))
		if answer == 1 :
			day_time_name = time_service(who_am_i)
			go_work_excel(day_time_name)
		elif answer == 2 :
			day_time_name = time_service(who_am_i)	
			leave_office_excel(day_time_name)
elif choice == "3" :
	sys.exit()
